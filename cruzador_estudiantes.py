# cruzador_estudiante.py
import io
import re
import hashlib
from pathlib import Path
from typing import Dict, Optional, Tuple, List
from openpyxl.styles import PatternFill

import numpy as np
import pandas as pd
import streamlit as st
import difflib


# =========================
# Config
# =========================
ROOT = Path(__file__).resolve().parent
PLANTILLAS_DIR = ROOT / "plantillas"

DEFAULT_AKADEMIC = ROOT / "alumnos-en akademic.xlsx"
DEFAULT_CURLLE = ROOT / "Data Historica Notas Curlle (1).csv"
DEFAULT_PLANES = ROOT / "todos los planes-juntos.xlsx"
DEFAULT_TODOS_PLANES_AKADEMIC = ROOT / "TODOS_PLANES_AKADEMIC.xlsx"

P01_TEMPLATE = PLANTILLAS_DIR / "P01_existentes_en_akademic.xlsx"
P02_TEMPLATE = PLANTILLAS_DIR / "P02_no_akademic_no_curlle.xlsx"
P03_TEMPLATE = PLANTILLAS_DIR / "P03_no_akademic_si_curlle.xlsx"
P04_TEMPLATE = PLANTILLAS_DIR / "P04_matricula_desde_akademic.xlsx"
P05_TEMPLATE = PLANTILLAS_DIR / "P05_carga_notas_akademic.xlsx"


# =========================
# Helpers base (BLINDADOS)
# =========================
def norm_col(c: str) -> str:
    """
    Normaliza encabezados para que:
    - Respete '_' (underscore)
    - Convierta '.', '-', '/', etc. en espacios
    - Quite tildes y caracteres raros
    - Devuelva snake_case consistente
    """
    s = "" if c is None else str(c)
    s = s.strip().lower()

    # separadores comunes a espacio
    s = re.sub(r"[.\-\/\\]+", " ", s)

    # espacios múltiples
    s = re.sub(r"\s+", " ", s)

    # quitar tildes
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )

    # dejar solo letras, números, espacio y underscore
    s = re.sub(r"[^a-z0-9 _]+", "", s)

    # snake_case
    s = s.replace(" ", "_")
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def norm_text_keep_spaces(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return s.strip()


def only_digits(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\D+", "", str(s))


def zfill8(s) -> str:
    d = only_digits(s)
    if not d:
        return ""
    return d.zfill(8)


def periodo_formato(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"\.0$", "", s)
    if re.fullmatch(r"\d{5}", s):  # 20201
        return f"{s[:4]}-{s[4:]}"
    return s


def normalize_cod_curso_spaces(x) -> str:
    """
    ✅ NUEVO: normaliza códigos como '10 A06' -> '10A06'
    - quita espacios internos (y cualquier whitespace)
    - upper
    """
    s = "" if x is None else str(x)
    s = s.strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """
    Busca una columna del df usando candidatos.
    - normaliza candidates
    - además intenta match "suave" quitando underscores
      (ej: codigo_estudiante == codigoestudiante)
    """
    cols = list(df.columns)
    cols_set = set(cols)
    soft_map = {c.replace("_", ""): c for c in cols}

    for cand in candidates:
        cn = norm_col(cand)
        if cn in cols_set:
            return cn
        cn_soft = cn.replace("_", "")
        if cn_soft in soft_map:
            return soft_map[cn_soft]

    return None


def build_fullname_from_parts(padron: pd.DataFrame, ap_col: str, am_col: str, nom_col: str) -> pd.Series:
    return (
        padron[ap_col].fillna("").astype(str).str.strip()
        + " "
        + padron[am_col].fillna("").astype(str).str.strip()
        + " "
        + padron[nom_col].fillna("").astype(str).str.strip()
    ).str.replace(r"\s+", " ", regex=True).str.strip()


# =========================
# ✅ LECTURA EXCEL ESTABLE (ANTI-CRASH)
# =========================
HEADER_HINTS = [
    "dni", "documento", "codigo_estudiante", "codigo_alumno", "codigo",
    "apellidos", "nombres", "apellido", "programa", "carrera",
    "plan", "plan_sigu", "periodo", "periodo_ingreso"
]


def _to_bytes(uploaded_or_path) -> bytes:
    if uploaded_or_path is None:
        return b""
    if hasattr(uploaded_or_path, "getvalue"):
        return uploaded_or_path.getvalue()
    if hasattr(uploaded_or_path, "read"):
        pos = uploaded_or_path.tell()
        b = uploaded_or_path.read()
        try:
            uploaded_or_path.seek(pos)
        except Exception:
            pass
        return b
    return Path(uploaded_or_path).read_bytes()


def _best_header_row_from_preview(preview_df: pd.DataFrame, max_rows: int = 60) -> int:
    best_i = 0
    best_score = -1
    hints = [norm_col(h) for h in HEADER_HINTS]

    for i in range(min(max_rows, len(preview_df))):
        row = preview_df.iloc[i].tolist()
        tokens = []
        for v in row:
            if v is None:
                continue
            s = str(v).strip()
            if s == "":
                continue
            tokens.append(norm_col(s))

        if not tokens:
            continue

        joined = " ".join(tokens)
        score = 0
        for hn in hints:
            if hn in tokens or hn in joined:
                score += 1

        non_empty = sum(1 for t in tokens if t != "")
        if non_empty >= 8:
            score += 2
        if non_empty >= 12:
            score += 2

        if score > best_score:
            best_score = score
            best_i = i

    if best_score < 2:
        return 0
    return best_i


def read_excel_any(file_or_path, sheet_name=0) -> pd.DataFrame:
    bio = io.BytesIO(_to_bytes(file_or_path))

    # preview (60 filas)
    try:
        bio.seek(0)
        preview = pd.read_excel(
            bio, dtype=str, engine="calamine",
            sheet_name=sheet_name, header=None, nrows=60
        )
    except Exception:
        bio.seek(0)
        preview = pd.read_excel(
            bio, dtype=str, engine="openpyxl",
            sheet_name=sheet_name, header=None, nrows=60
        )

    header_i = _best_header_row_from_preview(preview, max_rows=60)

    # lectura final
    try:
        bio.seek(0)
        df = pd.read_excel(
            bio, dtype=str, engine="calamine",
            sheet_name=sheet_name, header=0, skiprows=header_i
        )
    except Exception:
        bio.seek(0)
        df = pd.read_excel(
            bio, dtype=str, engine="openpyxl",
            sheet_name=sheet_name, header=0, skiprows=header_i
        )

    df.columns = [norm_col(c) for c in df.columns]
    return df


# =========================
# Curlle CSV (robusto)
# =========================
def find_header_row_in_curlle(csv_bytes_or_path) -> int:
    if hasattr(csv_bytes_or_path, "read"):
        raw = csv_bytes_or_path.read()
        csv_bytes_or_path.seek(0)
        lines = raw.splitlines()
    else:
        raw = Path(csv_bytes_or_path).read_bytes()
        lines = raw.splitlines()

    pats = [b"Codigo Alumno", "Código Alumno".encode("utf-8")]
    for i, line in enumerate(lines[:400]):
        if any(p in line for p in pats):
            return i
    return 3


def read_curlle(csv_file_or_path) -> pd.DataFrame:
    header_idx = find_header_row_in_curlle(csv_file_or_path)
    df = pd.read_csv(
        csv_file_or_path,
        dtype=str,
        sep=";",
        encoding="utf-8-sig",
        engine="python",
        skiprows=header_idx,
    )
    df.columns = [norm_col(c) for c in df.columns]
    return df


def template_columns(template_path: Path) -> Tuple[pd.DataFrame, str]:
    wb = pd.read_excel(template_path, sheet_name=None, dtype=str, engine="openpyxl")
    sheet = list(wb.keys())[0]
    df = wb[sheet].copy()
    df.columns = ["" if c is None else str(c) for c in df.columns]
    return df, sheet


def align_df_to_template_df(template_path: Path, data: pd.DataFrame) -> pd.DataFrame:
    tpl_df, _ = template_columns(template_path)
    tpl_cols = list(tpl_df.columns)

    tpl_map = {norm_col(c): c for c in tpl_cols}
    data_cols_map = {norm_col(c): c for c in data.columns}

    out = pd.DataFrame(columns=tpl_cols)
    for nkey, original_tpl_col in tpl_map.items():
        if nkey in data_cols_map:
            out[original_tpl_col] = data[data_cols_map[nkey]]
        else:
            out[original_tpl_col] = ""
    return out


def _sanitize_df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()

    out = df.copy()
    out.columns = ["" if c is None else str(c) for c in out.columns]
    out = out.replace({np.nan: "", pd.NaT: ""})

    for c in out.columns:
        out[c] = out[c].map(lambda v: "" if v is None else str(v))

    return out


def build_multi_sheet_excel(sheets: Dict[str, pd.DataFrame]) -> bytes:
    safe_sheets = {}
    for sheet_name, df in sheets.items():
        sh = "" if sheet_name is None else str(sheet_name)
        sh = sh.replace("/", "_").replace("\\", "_").replace("[", "(").replace("]", ")").replace("*", "_").replace("?", "_").replace(":", "_")
        sh = sh[:31] if sh else "Sheet"
        safe_sheets[sh] = _sanitize_df_for_excel(df)

    bio = io.BytesIO()

    try:
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            for sh, df in safe_sheets.items():
                df.to_excel(writer, index=False, sheet_name=sh)

                # ✅ Pintar amarillo filas con OBS en P03 (solo en Excel descargado)
                if sh == "P03" and ("OBS" in df.columns):
                    ws = writer.sheets[sh]

                    yellow_fill = PatternFill(
                        start_color="FFF59D",
                        end_color="FFF59D",
                        fill_type="solid",
                    )

                    # columna OBS (1-based en openpyxl)
                    obs_col_idx = list(df.columns).index("OBS") + 1

                    # recorrer filas de datos (row 1 es header, empieza en 2)
                    for row_idx in range(2, len(df) + 2):
                        obs_val = ws.cell(row=row_idx, column=obs_col_idx).value
                        if obs_val is not None and str(obs_val).strip() != "":
                            for col_idx in range(1, len(df.columns) + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

        return bio.getvalue()

    except Exception as e1:
        bio = io.BytesIO()
        try:
            # ⚠️ Fallback: xlsxwriter NO aplica estilos con openpyxl PatternFill
            # (pero sirve para no romper la descarga si openpyxl falla)
            with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
                for sh, df in safe_sheets.items():
                    df.to_excel(writer, index=False, sheet_name=sh)
            return bio.getvalue()
        except Exception as e2:
            raise RuntimeError(
                "No se pudo generar el Excel.\n"
                f"- Error con openpyxl: {type(e1).__name__}: {e1}\n"
                f"- Error con xlsxwriter: {type(e2).__name__}: {e2}\n\n"
                "Solución: instala xlsxwriter:\n"
                "   pip install -U xlsxwriter\n"
            )


def resolve_course_name_and_origin(pl2: pd.DataFrame, carrera_key: str, plan_key: str, cod_curso_key: str) -> Tuple[str, str, str]:
    carrera_key = "" if carrera_key is None else str(carrera_key).strip()
    plan_key = "" if plan_key is None else str(plan_key).strip()
    cod_curso_key = "" if cod_curso_key is None else str(cod_curso_key).strip().upper()

    if cod_curso_key == "":
        return "", carrera_key, plan_key

    exact = pl2[
        (pl2["carrera_key"] == carrera_key)
        & (pl2["plan_key"] == plan_key)
        & (pl2["cod_curso_key"] == cod_curso_key)
    ]
    if len(exact) > 0:
        row = exact.iloc[0]
        curso = str(row.get("curso", "")).strip()
        return curso, carrera_key, plan_key

    by_course = pl2[pl2["cod_curso_key"] == cod_curso_key].copy()
    if len(by_course) == 0:
        return "", carrera_key, plan_key

    by_course["curso"] = by_course["curso"].fillna("").astype(str).str.strip()
    cursos_no_vacios = [c for c in by_course["curso"].unique().tolist() if c != ""]
    cursos_no_vacios = list(dict.fromkeys(cursos_no_vacios))

    if len(cursos_no_vacios) == 1:
        chosen_curso = cursos_no_vacios[0]
        chosen_row = by_course[by_course["curso"] == chosen_curso].iloc[0]
        carrera_res = str(chosen_row.get("carrera_key", "")).strip()
        plan_res = str(chosen_row.get("plan_key", "")).strip()
        return chosen_curso, carrera_res, plan_res

    return "", carrera_key, plan_key


def coalesce_series(a: pd.Series, b: pd.Series) -> pd.Series:
    a = a.fillna("").astype(str)
    b = b.fillna("").astype(str)
    out = a.copy()
    mask = out.str.strip().eq("")
    out[mask] = b[mask]
    return out


# =========================
# ✅ Depuración FINAL P03 (CORREGIDA + REGLA REEMPLAZO)
# =========================
PROGRAMA_TO_COD = {
    norm_text_keep_spaces("INGENIERÍA INDUSTRIAL"): "IN",
    norm_text_keep_spaces("ADMINISTRACIÓN DE EMPRESAS"): "AE",
    norm_text_keep_spaces("INGENIERÍA CIVIL"): "IC",
    norm_text_keep_spaces("INGENIERÍA DE SISTEMAS"): "IS",
    norm_text_keep_spaces("ARQUITECTURA"): "AR",
    norm_text_keep_spaces("CONTABILIDAD"): "CO",
    norm_text_keep_spaces("DERECHO"): "DE",
    norm_text_keep_spaces("PSICOLOGÍA"): "PS",
    norm_text_keep_spaces("ENFERMERÍA"): "EN",
    norm_text_keep_spaces("MEDICINA HUMANA"): "MH",
    norm_text_keep_spaces("OBSTETRICIA"): "OB",
    norm_text_keep_spaces("ADMINISTRACIÓN Y FINANZA"): "AF",
}

FAM_ING = {"AF", "DE", "IS", "CO", "IN", "AE", "IC", "AR"}
FAM_SALUD = {"PS", "EN", "OB", "MH"}

COURSE_STOPWORDS = {
    "de", "del", "la", "las", "el", "los", "y", "e", "en", "para", "por", "a", "al",
    "un", "una", "unos", "unas", "i",
}

ROMAN_KEEP = {"ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x"}  # se conservan


def parse_nota_to_float(x) -> float:
    s = "" if x is None else str(x).strip()
    if s == "":
        return np.nan
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan


def course_base_name(cod_curso: str, nom_curso: str) -> str:
    cod = "" if cod_curso is None else str(cod_curso).strip().upper()
    nom = "" if nom_curso is None else str(nom_curso).strip()
    nom_n = norm_text_keep_spaces(nom)

    if "examen de suficiencia" in nom_n:
        nom_n = nom_n.replace("examen de suficiencia", "").strip()
        nom_n = re.sub(r"\s+", " ", nom_n).strip()

    nom_n = re.sub(r"\s*-\s*$", "", nom_n).strip()
    nom_n = re.sub(r"\s*-\s+", " ", nom_n).strip()

    return nom_n if nom_n != "" else cod


def compatible_family(student_cod: str) -> set:
    if not student_cod:
        return set()
    if student_cod in FAM_ING:
        return FAM_ING
    if student_cod in FAM_SALUD:
        return FAM_SALUD
    return {student_cod}


def _soft_singularize_token(w: str) -> str:
    if (
        len(w) >= 7 and w.endswith("s")
        and not w.endswith(("sis", "xis", "tis", "is"))
    ):
        return w[:-1]
    return w


def course_match_key(text: str) -> str:
    """
    Key fuerte para dedupe:
    - normaliza (tildes, etc.)
    - quita SOLO sufijo I/1 al final (NO toca II, III...)
    - elimina stopwords (de, y, del, ...)
    - singulariza suave
    - ordena tokens => "A y B" == "B A"
    - mantiene II/III/IV como token para que "X" != "X II"
    """
    t = norm_text_keep_spaces(text)

    # quitar SOLO ' i' o ' 1' al final
    t = re.sub(r"\s+(i|1)\s*$", "", t).strip()

    if t == "":
        return ""

    toks = [x for x in t.split() if x.strip() != ""]

    cleaned = []
    for w in toks:
        if w in COURSE_STOPWORDS:
            continue
        w = _soft_singularize_token(w)

        if w in ROMAN_KEEP:
            cleaned.append(w)
            continue

        if w.isdigit():
            cleaned.append(w)
            continue

        cleaned.append(w)

    if not cleaned:
        cleaned = [_soft_singularize_token(w) for w in toks if w not in COURSE_STOPWORDS]

    cleaned = [c for c in cleaned if c != ""]
    cleaned.sort()
    return " ".join(cleaned).strip()


REEMPLAZO_DESDE_PERIODO = "2018-1"


def _is_periodo_like(p: str) -> bool:
    return bool(re.fullmatch(r"\d{4}\-\d{1,2}", (p or "").strip()))


def _periodo_to_tuple(p: str):
    if not _is_periodo_like(p):
        return None
    y, t = p.split("-")
    try:
        return (int(y), int(t))
    except Exception:
        return None


def _period_ge(p: str, ref: str) -> bool:
    a = _periodo_to_tuple(p)
    b = _periodo_to_tuple(ref)
    if a is None or b is None:
        return False
    return a >= b


KEY_REDA_COMU = course_match_key("Redacción y Comunicación")
KEY_TALLER_ORAL = course_match_key("Taller de Comunicación Oral")
KEY_TALLER_ESCRITA = course_match_key("Taller de Comunicación Escrita")


def depurar_p03(
    si_df: pd.DataFrame,
    padron_prog_col: Optional[str],
    padron_plan_sigu_col: Optional[str],
    padron_cod_col: str
) -> pd.DataFrame:
    df = si_df.copy()
    df["alumno_key"] = df[padron_cod_col].fillna("").astype(str).str.strip()

    # Programa / Plan del alumno (padron)
    if padron_prog_col and padron_prog_col in df.columns:
        prog_norm = df[padron_prog_col].fillna("").astype(str).map(norm_text_keep_spaces)
        df["cod_programa_alumno"] = prog_norm.map(lambda x: PROGRAMA_TO_COD.get(x, "")).fillna("")
    else:
        df["cod_programa_alumno"] = ""

    if padron_plan_sigu_col and padron_plan_sigu_col in df.columns:
        df["plan_sigu_alumno"] = df[padron_plan_sigu_col].fillna("").astype(str).str.strip()
    else:
        df["plan_sigu_alumno"] = ""

    # Normalizaciones base
    df["nota_num"] = df["nota_curlle"].map(parse_nota_to_float)
    df["cod_curso_u"] = df["cod_curso"].fillna("").astype(str).map(normalize_cod_curso_spaces)
    df["nom_curso_u"] = df["curso_resuelto"].fillna("").astype(str).str.strip()

    df["curso_base"] = df.apply(lambda r: course_base_name(r["cod_curso_u"], r["nom_curso_u"]), axis=1)
    df["curso_match_key"] = df["curso_base"].map(course_match_key)
    df["periodo_u"] = df["periodo_fmt"].fillna("").astype(str).str.strip()

    # fallback si quedó vacío
    df.loc[df["curso_match_key"].eq(""), "curso_match_key"] = df["cod_curso_u"].fillna("").astype(str).str.strip()

    # ✅ CAMBIO 1 (HISTÓRICO):
    # Si en alumno+periodo+curso hay alguna nota >0 => eliminar notas 0 del MISMO PERIODO (no cruza años)
    df["_nota_safe"] = df["nota_num"].where(~df["nota_num"].isna(), -1)

    has_real_positive = (
        (df["_nota_safe"] > 0)
        .groupby([df["alumno_key"], df["periodo_u"], df["curso_match_key"]])
        .any()
        .to_dict()
    )

    drop_zero = []
    for idx, row in df.iterrows():
        key = (row["alumno_key"], row["periodo_u"], row["curso_match_key"])
        cond_has_real = bool(has_real_positive.get(key, False))
        cond_is_zero = (row["_nota_safe"] == 0)
        drop_zero.append(cond_has_real and cond_is_zero)

    df = df.loc[~pd.Series(drop_zero, index=df.index)].copy()

    # -------------------------------------------------------
    # ✅ REEMPLAZO: Redacción y Comunicación reemplaza Taller Oral/Escrita
    # -------------------------------------------------------
    in_scope = df["periodo_u"].map(lambda p: _period_ge(p, REEMPLAZO_DESDE_PERIODO))

    has_reda = (
        df.loc[in_scope, ["alumno_key", "curso_match_key"]]
          .assign(is_reda=lambda x: x["curso_match_key"].eq(KEY_REDA_COMU))
          .groupby("alumno_key")["is_reda"]
          .any()
          .to_dict()
    )

    drop_rep = []
    for idx, row in df.iterrows():
        alumno = row["alumno_key"]
        ck = row["curso_match_key"]
        per = row["periodo_u"]

        if not _period_ge(per, REEMPLAZO_DESDE_PERIODO):
            drop_rep.append(False)
            continue

        if bool(has_reda.get(alumno, False)) and ck in {KEY_TALLER_ORAL, KEY_TALLER_ESCRITA}:
            drop_rep.append(True)
        else:
            drop_rep.append(False)

    df = df.loc[~pd.Series(drop_rep, index=df.index)].copy()

    # Scoring: preferir plan/carrera que calcen con padron
    plan_out = df["plan_out"].fillna("").astype(str).str.strip()
    carrera_out = df["cod_carrera_out"].fillna("").astype(str).str.strip()
    plan_sigu = df["plan_sigu_alumno"].fillna("").astype(str).str.strip()
    cod_prog = df["cod_programa_alumno"].fillna("").astype(str).str.strip()

    score = pd.Series(0, index=df.index, dtype="int64")

    m_plan = (plan_sigu != "") & plan_out.eq(plan_sigu)
    score.loc[m_plan] += 4

    m_carr = (cod_prog != "") & carrera_out.eq(cod_prog)
    score.loc[m_carr] += 2

    fam_ok = []
    for i in df.index:
        fam = compatible_family(cod_prog.loc[i])
        fam_ok.append((carrera_out.loc[i] in fam) if fam else False)
    fam_ok = pd.Series(fam_ok, index=df.index)
    score.loc[(~m_carr) & fam_ok] += 1

    df["_score"] = score
    df["_idx0"] = np.arange(len(df))

    # ✅ DEDUPE FINAL
    group_cols = ["alumno_key", "periodo_u", "curso_match_key"]

    df = (
        df.sort_values(
            by=group_cols + ["_score", "_nota_safe", "_idx0"],
            ascending=[True, True, True, False, False, True],
        )
        .drop_duplicates(subset=group_cols, keep="first")
        .drop(columns=["_score", "_idx0", "_nota_safe"])
        .copy()
    )

    return df


# =========================
# ✅ P04 (Matricula) - regla por matriz ACTIVO/INACTIVO
# =========================
PROG_TO_ESCUELA = {
    norm_text_keep_spaces("INGENIERÍA INDUSTRIAL"): "P06",
    norm_text_keep_spaces("ADMINISTRACIÓN DE EMPRESAS"): "P38",
    norm_text_keep_spaces("ADMINISTRACIÓN DE EMPRESAS - VIRTUAL"): "P38V",
    norm_text_keep_spaces("INGENIERÍA CIVIL"): "P32",
    norm_text_keep_spaces("INGENIERÍA DE SISTEMAS"): "P03",
    norm_text_keep_spaces("ARQUITECTURA"): "P34",
    norm_text_keep_spaces("CONTABILIDAD"): "P07",
    norm_text_keep_spaces("CONTABILIDAD - VIRTUAL"): "P07V",
    norm_text_keep_spaces("DERECHO"): "P08",
    norm_text_keep_spaces("DERECHO - VIRTUAL"): "P08V",
    norm_text_keep_spaces("PSICOLOGÍA"): "P04",
    norm_text_keep_spaces("ENFERMERÍA"): "P02",
    norm_text_keep_spaces("MEDICINA HUMANA"): "P31",
    norm_text_keep_spaces("OBSTETRICIA"): "P09",
    norm_text_keep_spaces("TECNOLOGÍA MÉDICA - ESPECIALIDAD EN TERAPIA DE LENGUAJE"): "P37",
    norm_text_keep_spaces("TECNOLOGÍA MÉDICA - ESPECIALIDAD EN TERAPIA FÍSICA Y REHABILITACIÓN"): "P36",
    norm_text_keep_spaces("TECNOLOGÍA MÉDICA - OPTOMETRÍA"): "P55",
    norm_text_keep_spaces("TECNOLOGÍA MÉDICA - ESPECIALIDAD EN LABORATORIO CLINICO Y ANATOMÍA PATOLÓGICA"): "P35",
    norm_text_keep_spaces("ADMINISTRACIÓN Y FINANZA"): "P01",
    norm_text_keep_spaces("ADMINISTRACION Y FINANZAS"): "P01",
    norm_text_keep_spaces("ADMINISTRACION Y FINANZA"): "P01",
}

PLAN_ORDER = ["14210", "20172", "201722", "20201", "202011", "20222", "202312", "20242", "20251"]

ESCUELA_PLANES_ACTIVOS = {
    "P01": {"202011", "20242"},
    "P03": {"202011", "20242"},
    "P05": set(),
    "P06": {"202011", "20242"},
    "P32": {"20242"},
    "P08": {"201722", "202312", "20242", "20251"},
    "P07": {"201722", "20242"},
    "P02": {"202011", "20242"},
    "P04": {"202011", "20242"},
    "P31": {"20242", "20251"},
    "P09": {"201722", "20242"},
}


def escuela_from_programa(programa: str) -> str:
    key = norm_text_keep_spaces(programa)
    return PROG_TO_ESCUELA.get(key, "")


def tipo_matricula_from_codcurso(cod_curso: str) -> str:
    c = normalize_cod_curso_spaces(cod_curso)
    return "EXSUF" if c.startswith("EXSUF") else "R"


def _plan_to_num(s: str) -> Optional[int]:
    t = "" if s is None else str(s).strip()
    if t == "":
        return None
    t = re.sub(r"\D+", "", t)
    if t == "":
        return None
    try:
        return int(t)
    except Exception:
        return None


def normalize_plan_by_matrix(plan_sigu: str, escuela: str) -> str:
    p = "" if plan_sigu is None else str(plan_sigu).strip()
    if p == "":
        return ""
    if escuela not in ESCUELA_PLANES_ACTIVOS:
        return p

    active_set = ESCUELA_PLANES_ACTIVOS.get(escuela, set())
    if not PLAN_ORDER:
        return ""

    pnum = _plan_to_num(p)
    if pnum is None:
        return ""

    order_nums = [int(x) for x in PLAN_ORDER]
    start_idx = None
    for i, xnum in enumerate(order_nums):
        if xnum >= pnum:
            start_idx = i
            break
    if start_idx is None:
        return ""

    for j in range(start_idx, len(PLAN_ORDER)):
        candidate = PLAN_ORDER[j]
        if candidate in active_set:
            return candidate

    return ""


def is_egresado_flag(val, codigo_alumno=None) -> bool:
    s = "" if val is None else str(val).strip()
    if s == "":
        return False

    su = s.upper()
    if su in {"#N/D", "#N/A", "N/A", "NA"}:
        return False

    cod = "" if codigo_alumno is None else str(codigo_alumno).strip()
    if cod == "":
        return False

    return only_digits(s) != "" and only_digits(s) == only_digits(cod)


# =========================
# ✅ CodigoCurso (P04) desde TODOS_PLANES_AKADEMIC
# =========================
def _clean_codcurso(x: str) -> str:
    s = "" if x is None else str(x).strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def _course_name_key(name: str) -> str:
    """
    Normaliza nombre de curso para matchear contra TODOS_PLANES_AKADEMIC:
    - quita 'examen de suficiencia'
    - elimina stopwords comunes
    - singulariza suave
    - ordena tokens
    """
    n0 = "" if name is None else str(name)
    n = norm_text_keep_spaces(n0)

    if n == "":
        return ""

    n = n.replace("examen de suficiencia", " ").strip()
    n = re.sub(r"\s+", " ", n).strip()
    n = re.sub(r"\s*-\s*$", "", n).strip()

    stop = set(COURSE_STOPWORDS) | {"examen", "suficiencia"}
    toks = [t for t in n.split() if t and t not in stop]

    toks = [_soft_singularize_token(t) for t in toks]
    toks = [t for t in toks if t]

    toks.sort()
    return " ".join(toks).strip()


def build_todos_planes_akademic_map(df_tpa: pd.DataFrame) -> pd.DataFrame:
    if "codigo" not in df_tpa.columns:
        raise RuntimeError("TODOS_PLANES_AKADEMIC no trae la columna 'CODIGO'.")

    curso_col = "curso" if "curso" in df_tpa.columns else None
    if not curso_col:
        raise RuntimeError("TODOS_PLANES_AKADEMIC no trae la columna 'CURSO' (necesaria para el match por nombre).")

    t = df_tpa.copy()
    t["codigo_raw"] = t["codigo"].fillna("").astype(str).str.strip()
    t["curso_raw"] = t[curso_col].fillna("").astype(str).str.strip()

    codigo_clean = t["codigo_raw"].map(lambda x: re.sub(r"\s+", "", str(x))).str.strip()

    m = codigo_clean.str.extract(
        r"^(?P<escuela>P\d{2,3}[A-Z]?)\-(?P<plan>\d{3,10})\-(?P<curso>.+)$",
        expand=True,
    )

    t["escuela_key"] = m["escuela"].fillna("").astype(str).str.strip()
    t["plan_key"] = m["plan"].fillna("").astype(str).str.strip()
    t["codcurso_key"] = m["curso"].fillna("").astype(str).map(_clean_codcurso)
    t["curso_name_key"] = t["curso_raw"].map(_course_name_key)

    out = t[["codigo_raw", "escuela_key", "plan_key", "codcurso_key", "curso_raw", "curso_name_key"]].copy()
    out = out[out["codigo_raw"].astype(str).str.strip() != ""].copy()
    return out


# =========================================================
# ✅ BLINDAJE: evitar fuzzy-match incorrecto con niveles (II/III/IV)
# =========================================================
ROMAN_LEVELS = {"i", "ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x"}


def _extract_roman_level(name: str) -> str:
    """
    Devuelve el nivel romano encontrado (i/ii/iii/iv...) si está como token.
    Si no hay, devuelve "".
    """
    t = norm_text_keep_spaces(name)
    if not t:
        return ""
    toks = t.split()
    if not toks:
        return ""
    last = toks[-1]
    if last in ROMAN_LEVELS:
        return last
    for w in reversed(toks):
        if w in ROMAN_LEVELS:
            return w
    return ""


def _roman_level_compatible(target_name: str, cand_name: str) -> bool:
    """
    Reglas:
    - Si target tiene nivel (II/III/IV...), el candidato DEBE tener el MISMO.
    - Si target NO tiene nivel, el candidato tampoco debe tener.
    """
    lt = _extract_roman_level(target_name)
    lc = _extract_roman_level(cand_name)
    if lt:
        return lc == lt
    return lc == ""


def _best_fuzzy_match_code(same_plan_df: pd.DataFrame, nombre_curso: str, min_ratio: float = 0.86) -> str:
    """
    Mejor match por parecido (difflib) dentro del MISMO plan.
    ✅ BLINDADO: NO permite match si el nivel romano no coincide (II/III/IV...).
    Devuelve codigo_raw si supera umbral.
    """
    target_raw = "" if nombre_curso is None else str(nombre_curso)
    target = norm_text_keep_spaces(target_raw)
    if target == "":
        return ""

    best_ratio = 0.0
    best_code = ""

    for _, r in same_plan_df.iterrows():
        cand_raw = "" if r.get("curso_raw", "") is None else str(r.get("curso_raw", ""))
        cand = norm_text_keep_spaces(cand_raw)
        if cand == "":
            continue

        # ✅ BLINDAJE: nivel romano debe ser compatible
        if not _roman_level_compatible(target_raw, cand_raw):
            continue

        ratio = difflib.SequenceMatcher(None, target, cand).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_code = str(r.get("codigo_raw", "")).strip()

    if best_ratio >= min_ratio:
        return best_code
    return ""


def lookup_codigo_curso_tpa(tpa_map: pd.DataFrame, escuela: str, plan: str, codcurso: str, nombre_curso: str) -> str:
    """
    Regla:
    - Si NO es EXSUF: intenta match exacto por codcurso_key; si falla, por nombre key; si falla, fuzzy.
    - Si ES EXSUF: NO intentes por código (EXSUFxxIN no existe en TPA como codcurso_key),
      intenta por nombre (key) y luego fuzzy.
    """
    e = "" if escuela is None else str(escuela).strip()
    p = "" if plan is None else str(plan).strip()

    c_raw = normalize_cod_curso_spaces(codcurso)
    c = _clean_codcurso(c_raw)

    if e == "" or p == "":
        return ""

    same_plan = tpa_map[(tpa_map["escuela_key"] == e) & (tpa_map["plan_key"] == p)]
    if len(same_plan) == 0:
        return ""

    is_exsuf = c_raw.startswith("EXSUF")

    # 1) exact por código (solo si NO es EXSUF)
    if (not is_exsuf) and c != "":
        exact = same_plan[same_plan["codcurso_key"] == c]
        if len(exact) > 0:
            return str(exact.iloc[0]["codigo_raw"]).strip()

    # 2) exact por nombre normalizado (key)
    nk = _course_name_key(nombre_curso)
    if nk != "":
        same_plan2 = same_plan[same_plan["curso_name_key"] == nk]
        if len(same_plan2) > 0:
            return str(same_plan2.iloc[0]["codigo_raw"]).strip()

    # 3) fuzzy por parecido dentro del MISMO plan (blindado por nivel romano)
    fb = _best_fuzzy_match_code(same_plan, nombre_curso=nombre_curso, min_ratio=0.86)
    if fb:
        return fb

    return ""


# =========================
# ✅ Cache robusto (por HASH)
# =========================
def _hash_bytes(b: bytes) -> str:
    return hashlib.md5(b).hexdigest()


@st.cache_data(show_spinner=False)
def _cached_read_excel_bytes(file_bytes: bytes, _md5: str) -> pd.DataFrame:
    return read_excel_any(io.BytesIO(file_bytes))


@st.cache_data(show_spinner=False)
def _cached_read_excel_path(path_str: str) -> pd.DataFrame:
    return read_excel_any(path_str)


@st.cache_data(show_spinner=False)
def _cached_read_curlle_bytes(file_bytes: bytes, _md5: str) -> pd.DataFrame:
    return read_curlle(io.BytesIO(file_bytes))


@st.cache_data(show_spinner=False)
def _cached_read_curlle_path(path_str: str) -> pd.DataFrame:
    return read_curlle(path_str)


# =========================
# UI
# =========================
st.set_page_config(page_title="Regularización AKA", layout="wide")
st.title("📌 Regularización AKA - Generador de Plantillas")

st.markdown(
    """
Sube el **padrón principal** (archivo con alumnos).
La app cruza contra:
- **Akademic** (existentes) por **DNI o Código**
- **Curlle** (histórica de notas) por **Código Alumno**
- **Planes** (todos los planes-juntos) para obtener **Nom. Curso**
- **TODOS_PLANES_AKADEMIC** para armar **CodigoCurso** correcto en P04
"""
)

with st.sidebar:
    st.header("📥 Cargas")

    st.subheader("📌 Base detectada en la raíz")
    st.caption(f"Carpeta actual: {ROOT}")

    ok_ak = DEFAULT_AKADEMIC.exists()
    ok_cu = DEFAULT_CURLLE.exists()
    ok_pl = DEFAULT_PLANES.exists()
    ok_tpa = DEFAULT_TODOS_PLANES_AKADEMIC.exists()

    st.write(("✅" if ok_ak else "❌"), DEFAULT_AKADEMIC.name)
    st.write(("✅" if ok_cu else "❌"), DEFAULT_CURLLE.name)
    st.write(("✅" if ok_pl else "❌"), DEFAULT_PLANES.name)
    st.write(("✅" if ok_tpa else "❌"), DEFAULT_TODOS_PLANES_AKADEMIC.name)

    st.divider()
    padron_file = st.file_uploader("1) Padrón (OBLIGATORIO) - Excel/CSV", type=["xlsx", "xls", "csv"])

    st.divider()
    usar_override = st.checkbox("Quiero reemplazar archivos base (override)", value=False)

    akademic_file = None
    curlle_file = None
    planes_file = None
    todos_planes_ak_file = None

    if usar_override:
        akademic_file = st.file_uploader("Akademic (xlsx)", type=["xlsx", "xls"])
        curlle_file = st.file_uploader("Curlle (csv)", type=["csv"])
        planes_file = st.file_uploader("Planes (xlsx)", type=["xlsx", "xls"])
        todos_planes_ak_file = st.file_uploader("TODOS_PLANES_AKADEMIC (xlsx)", type=["xlsx", "xls"])

    st.divider()
    run_btn = st.button("🚀 Generar plantillas", type="primary")

if not padron_file:
    st.info("Sube el **padrón** para empezar.")
    st.stop()

if not run_btn:
    st.stop()

if not akademic_file and not DEFAULT_AKADEMIC.exists():
    st.error(f"No encuentro {DEFAULT_AKADEMIC.name} en la raíz. Marca override y súbelo.")
    st.stop()
if not curlle_file and not DEFAULT_CURLLE.exists():
    st.error(f"No encuentro {DEFAULT_CURLLE.name} en la raíz. Marca override y súbelo.")
    st.stop()
if not planes_file and not DEFAULT_PLANES.exists():
    st.error(f"No encuentro {DEFAULT_PLANES.name} en la raíz. Marca override y súbelo.")
    st.stop()
if not todos_planes_ak_file and not DEFAULT_TODOS_PLANES_AKADEMIC.exists():
    st.error(f"No encuentro {DEFAULT_TODOS_PLANES_AKADEMIC.name} en la raíz. Marca override y súbelo.")
    st.stop()

# =========================
# ✅ Guard por tamaño (evita tumbar Streamlit)
# =========================
MAX_MB = 35
if padron_file and padron_file.name.lower().endswith((".xlsx", ".xls")):
    size_mb = len(padron_file.getvalue()) / (1024 * 1024)
    if size_mb > MAX_MB:
        st.error(f"Tu padrón pesa {size_mb:.1f} MB. Recomendado < {MAX_MB} MB o súbelo como CSV.")
        st.stop()

# =========================
# Load Padrón (estable)
# =========================
with st.spinner("Leyendo PADRÓN..."):
    if padron_file.name.lower().endswith(".csv"):
        padron = pd.read_csv(padron_file, dtype=str, encoding="utf-8-sig")
        padron.columns = [norm_col(c) for c in padron.columns]
    else:
        padron_bytes = padron_file.getvalue()
        padron = _cached_read_excel_bytes(padron_bytes, _hash_bytes(padron_bytes))

with st.expander("🧪 Debug padrón: columnas detectadas"):
    st.write(padron.columns.tolist()[:120])

# =========================
# Detección columnas PADRÓN
# =========================
dni_candidates = [
    "dni", "documento", "doc", "document", "nro_documento", "n_documento",
    "numero_documento", "num_documento", "nro_doc", "numero_doc", "num_doc",
    "doc_identidad", "documento_identidad", "numero_de_documento", "nro_de_documento",
    "n_de_documento", "n_documento_identidad", "ndocumento"
]
cod_candidates = [
    "codigo_estudiante", "codigo_alumno", "codigo", "cod_alumno", "cod_estudiante",
    "cod_est", "codigo_est", "codigo_de_estudiante", "codigo_de_alumno",
    "codigo_estudiantil", "cod_estudiantil", "codigo_unico",
    "userusername", "username", "codigo_de_usuario", "user_username"
]
prog_candidates = [
    "programa_academico", "programa", "programa_academ", "academic_program",
    "carrera", "career", "careername",
    "programa_academico_nombre", "nombre_programa", "escuela", "especialidad"
]
ap_pat_candidates = ["apellido_paterno", "apellidopaterno", "ape_paterno", "ape_pat", "paterno", "apellido_pat", "ap_paterno"]
ap_mat_candidates = ["apellido_materno", "apellidomaterno", "ape_materno", "ape_mat", "materno", "apellido_mat", "am_materno", "ap_materno"]
nombres_candidates = ["nombres", "nombres_estudiante", "nombre", "name", "primer_nombre", "nombres_completos"]
egresados_candidates = ["egresados", "egresado", "egresada", "egresadas", "egresado_s", "egresado_flag"]
periodo_ingreso_candidates = [
    "periodo_ingreso", "periodo_de_ingreso", "periododeingreso",
    "ingreso", "per_ingreso", "per_ing", "periodoingreso",
    "periodo_academico_ingreso", "periodoacademico_ingreso"
]

padron_dni_col = pick_col(padron, dni_candidates)
padron_cod_col = pick_col(padron, cod_candidates)
padron_prog_col = pick_col(padron, prog_candidates)
ap_pat_col = pick_col(padron, ap_pat_candidates)
ap_mat_col = pick_col(padron, ap_mat_candidates)
nombres_col = pick_col(padron, nombres_candidates)
padron_egresados_col = pick_col(padron, egresados_candidates)
padron_periodo_ingreso_col = pick_col(padron, periodo_ingreso_candidates)

st.subheader("🔎 Mapeo de columnas del padrón (automático + fallback)")
st.write("DNI:", padron_dni_col, " | COD:", padron_cod_col, " | PROGRAMA:", padron_prog_col)

if not padron_dni_col:
    st.warning("No pude detectar DNI automáticamente. Selecciónalo manualmente.")
    padron_dni_col = st.selectbox("Selecciona la columna DNI", options=padron.columns.tolist())

if not padron_cod_col:
    st.warning("No pude detectar CODIGO_ESTUDIANTE automáticamente. Selecciónalo manualmente.")
    padron_cod_col = st.selectbox("Selecciona la columna CODIGO_ESTUDIANTE", options=padron.columns.tolist())

# Nombre completo
if ap_pat_col and ap_mat_col and nombres_col:
    padron["nombre_completo_calc"] = build_fullname_from_parts(padron, ap_pat_col, ap_mat_col, nombres_col)
else:
    nc = pick_col(
        padron,
        ["apellidos_y_nombres", "apellidos_nombres", "apellido_y_nombres",
         "nombre_completo", "fullname", "full_name", "nombres_completos"]
    )
    padron["nombre_completo_calc"] = padron[nc].fillna("").astype(str).str.strip() if nc else ""

# PLAN-SIGU
padron_plan_sigu_col = pick_col(
    padron,
    ["plansigu", "plan_sigu", "plan_sigui", "plan_siguiente", "plan_sgte",
     "plan_sig", "plan", "plan_de_estudios", "plan_estudios"]
)

# EGRESADO
if padron_egresados_col:
    padron["egresado_flag"] = padron.apply(
        lambda r: is_egresado_flag(r.get(padron_egresados_col), r.get(padron_cod_col)),
        axis=1
    )
else:
    padron["egresado_flag"] = False

# PERIODO INGRESO
if padron_periodo_ingreso_col:
    padron["periodo_ingreso_fmt"] = padron[padron_periodo_ingreso_col].map(periodo_formato)
else:
    padron["periodo_ingreso_fmt"] = ""

# =========================
# Load bases (estable)
# =========================
with st.spinner("Leyendo bases (Akademic / Planes / Todos Planes / Curlle)..."):

    if akademic_file:
        b = akademic_file.getvalue()
        ak = _cached_read_excel_bytes(b, _hash_bytes(b))
    else:
        ak = _cached_read_excel_path(str(DEFAULT_AKADEMIC))

    if planes_file:
        b = planes_file.getvalue()
        pl = _cached_read_excel_bytes(b, _hash_bytes(b))
    else:
        pl = _cached_read_excel_path(str(DEFAULT_PLANES))

    if todos_planes_ak_file:
        b = todos_planes_ak_file.getvalue()
        tpa = _cached_read_excel_bytes(b, _hash_bytes(b))
    else:
        tpa = _cached_read_excel_path(str(DEFAULT_TODOS_PLANES_AKADEMIC))

    if curlle_file:
        b = curlle_file.getvalue()
        cur = _cached_read_curlle_bytes(b, _hash_bytes(b))
    else:
        cur = _cached_read_curlle_path(str(DEFAULT_CURLLE))

try:
    tpa_map = build_todos_planes_akademic_map(tpa)
except Exception as e:
    st.error(str(e))
    st.stop()

# =========================
# Normalize keys
# =========================
padron["dni_key"] = padron[padron_dni_col].map(zfill8)
padron["cod_key"] = padron[padron_cod_col].fillna("").astype(str).str.strip().str.upper()

# Akademic: DNI y COD
dni_cols_ak = [c for c in ["user_dni", "userdni", "dni", "document", "documento", "numero_documento", "nro_documento"] if c in ak.columns]
ak["dni_key"] = ""
for c in dni_cols_ak:
    ak["dni_key"] = ak["dni_key"].where(ak["dni_key"].astype(str).str.len() > 0, ak[c].fillna(""))
ak["dni_key"] = ak["dni_key"].map(zfill8)

ak_dni_raw_col = pick_col(ak, ["user_dni", "userdni", "dni", "document", "documento", "numero_documento", "nro_documento"])
ak["dni_raw"] = ak[ak_dni_raw_col].fillna("").astype(str).str.strip() if ak_dni_raw_col else ""

ak_codigo_col = pick_col(ak, ["userusername", "username", "codigo_estudiante", "codigo_alumno", "cod_key"])
ak["cod_key"] = ""
if ak_codigo_col:
    ak["cod_key"] = ak[ak_codigo_col].fillna("").astype(str).str.strip().str.upper()

ak_prog_col = pick_col(ak, ["careername"])
ak_fullname_col = pick_col(ak, ["userfullname", "fullname"])
ak_plan_col = pick_col(ak, ["plan_de_estudios", "plan_de_estudio", "plan", "plan_estudios", "planestudios", "studyplan", "study_plan", "nombre_plan"])

ak_by_dni = ak[ak["dni_key"].fillna("").astype(str).str.len() > 0].drop_duplicates(subset=["dni_key"], keep="first")
ak_by_cod = ak[ak["cod_key"].fillna("").astype(str).str.len() > 0].drop_duplicates(subset=["cod_key"], keep="first")

# Curlle columnas requeridas
need_cols = ["codigo_alumno", "periodo", "cod_curso", "nota_curlle", "plan", "cod_carrera"]
for nc in need_cols:
    if nc not in cur.columns:
        st.error(f"Curlle no trae la columna esperada: {nc}")
        st.stop()

cur["cod_key"] = cur["codigo_alumno"].fillna("").astype(str).str.strip().str.upper()
cur["periodo_fmt"] = cur["periodo"].map(periodo_formato)

# ✅ CAMBIO 2 (ESPACIOS EN COD CURSO): normalizar '10 A06' -> '10A06'
cur["cod_curso"] = cur["cod_curso"].map(normalize_cod_curso_spaces)

cur["cod_curso_key"] = cur["cod_curso"].fillna("").astype(str).str.strip().str.upper()
cur["plan_key"] = cur["plan"].fillna("").astype(str).str.strip()
cur["carrera_key"] = cur["cod_carrera"].fillna("").astype(str).str.strip()

# Planes columnas requeridas
for nc in ["archivo", "codigo", "curso"]:
    if nc not in pl.columns:
        st.error(f"Planes no trae la columna esperada: {nc}")
        st.stop()

pl2 = pl.copy()
pl2["archivo"] = pl2["archivo"].fillna("").astype(str).str.strip()
ex = pl2["archivo"].str.extract(r"^(?P<carrera>[A-Za-z]{1,6})-(?P<plan>\d{3,10})", expand=True)
pl2["carrera_key"] = ex["carrera"].fillna("").astype(str).str.strip()
pl2["plan_key"] = ex["plan"].fillna("").astype(str).str.strip()
pl2["cod_curso_key"] = pl2["codigo"].fillna("").astype(str).str.strip().str.upper().map(normalize_cod_curso_spaces)
pl2["curso"] = pl2["curso"].fillna("").astype(str).str.strip()

# =========================
# 1) Cruce con Akademic (P01) por DNI O COD
# =========================
m1 = padron.merge(ak_by_dni, how="left", on="dni_key", suffixes=("", "_akdni"))
m2 = m1.merge(ak_by_cod, how="left", on="cod_key", suffixes=("", "_akcod"))

ak_code_dni = m2[ak_codigo_col] if ak_codigo_col and ak_codigo_col in m2.columns else pd.Series([""] * len(m2))
ak_code_cod = m2.get(f"{ak_codigo_col}_akcod", pd.Series([""] * len(m2))) if ak_codigo_col else pd.Series([""] * len(m2))

ak_prog_dni = m2[ak_prog_col] if ak_prog_col and ak_prog_col in m2.columns else pd.Series([""] * len(m2))
ak_prog_cod = m2.get(f"{ak_prog_col}_akcod", pd.Series([""] * len(m2))) if ak_prog_col else pd.Series([""] * len(m2))

ak_full_dni = m2[ak_fullname_col] if ak_fullname_col and ak_fullname_col in m2.columns else pd.Series([""] * len(m2))
ak_full_cod = m2.get(f"{ak_fullname_col}_akcod", pd.Series([""] * len(m2))) if ak_fullname_col else pd.Series([""] * len(m2))

ak_plan_dni = m2[ak_plan_col] if ak_plan_col and ak_plan_col in m2.columns else pd.Series([""] * len(m2))
ak_plan_cod = m2.get(f"{ak_plan_col}_akcod", pd.Series([""] * len(m2))) if ak_plan_col else pd.Series([""] * len(m2))

ak_dni_raw_dni = m2["dni_raw"] if "dni_raw" in m2.columns else pd.Series([""] * len(m2))
ak_dni_raw_cod = m2.get("dni_raw_akcod", pd.Series([""] * len(m2)))

m2["ak_codigo_final"] = coalesce_series(ak_code_dni, ak_code_cod).astype(str).str.strip()
m2["ak_prog_final"] = coalesce_series(ak_prog_dni, ak_prog_cod).astype(str).str.strip()
m2["ak_full_final"] = coalesce_series(ak_full_dni, ak_full_cod).astype(str).str.strip()
m2["ak_plan_final"] = coalesce_series(ak_plan_dni, ak_plan_cod).astype(str).str.strip()
m2["ak_dni_raw_final"] = coalesce_series(ak_dni_raw_dni, ak_dni_raw_cod).astype(str).str.strip()

exists_mask = (m2["ak_codigo_final"].str.strip() != "") | (m2["ak_full_final"].str.strip() != "")
en_ak = m2.loc[exists_mask].copy()
no_ak = m2.loc[~exists_mask].copy()

st.info(f"DEBUG: Padron={len(padron):,} | EnAkademic(P01)={len(en_ak):,} | NoAkademic={len(no_ak):,}")

p01_data = pd.DataFrame({
    "CODIGO_ESTUDIANTE": en_ak["ak_codigo_final"],
    "DNI": en_ak["ak_dni_raw_final"],
    "PROGRAMA ACADEMICO": en_ak["ak_prog_final"],
    "NOMBRES-COMPLETOS": en_ak["ak_full_final"],
    "PLAN-AKADEMIC": en_ak["ak_plan_final"],
}).drop_duplicates()

# =========================
# 2) Cruce no akademic vs Curlle (P02 / P03)
# =========================
no_ak_basic = no_ak.copy()
no_ak_basic["nombre_completo"] = no_ak_basic["nombre_completo_calc"]

cur_codcurso_clean = cur["cod_curso"].fillna("").astype(str).str.strip()
cur_valid_student_keys = set(cur.loc[cur_codcurso_clean.ne(""), "cod_key"].astype(str).str.strip().str.upper().unique())

no_ak_basic["has_curlle_student"] = no_ak_basic["cod_key"].fillna("").astype(str).str.strip().str.upper().isin(cur_valid_student_keys)
no_curlle = no_ak_basic.loc[~no_ak_basic["has_curlle_student"]].copy()

m_cur = no_ak_basic.loc[no_ak_basic["has_curlle_student"]].merge(
    cur,
    how="left",
    on="cod_key",
    suffixes=("", "_cur"),
)

m_cur["cod_curso_clean"] = m_cur["cod_curso"].fillna("").astype(str).str.strip()
si_curlle = m_cur.loc[m_cur["cod_curso_clean"].ne("")].copy()

st.info(
    "DEBUG CURLLE MATCH (ANTI-JOIN) -> "
    f"NoAkademic alumnos={len(no_ak_basic):,} | "
    f"ConCurlle(alumno)={int(no_ak_basic['has_curlle_student'].sum()):,} | "
    f"SinCurlle(P02 alumnos)={len(no_curlle):,} | "
    f"FilasCurlleReales(P03 base)={len(si_curlle):,}"
)

p02_data = pd.DataFrame({
    "CODIGO_ESTUDIANTE": no_curlle[padron_cod_col].fillna("").astype(str).str.strip(),
    "PROGRAMA ACADEMICO": no_curlle[padron_prog_col].fillna("").astype(str).str.strip() if padron_prog_col else "",
    "NOMBRES-COMPLETOS": no_curlle["nombre_completo"],
    "PLAN-SUBIDO": no_curlle[padron_plan_sigu_col].fillna("").astype(str).str.strip() if padron_plan_sigu_col else "",
})

# =========================
# P03: resolver nombre + origen (plan/carrera)
# =========================
si_curlle["carrera_key"] = si_curlle["carrera_key"].fillna("").astype(str).str.strip()
si_curlle["plan_key"] = si_curlle["plan_key"].fillna("").astype(str).str.strip()
si_curlle["cod_curso_key"] = si_curlle["cod_curso_key"].fillna("").astype(str).map(normalize_cod_curso_spaces)

resolved_curso, resolved_carrera, resolved_plan = [], [], []
for _, row in si_curlle.iterrows():
    curso, car_res, plan_res = resolve_course_name_and_origin(
        pl2=pl2,
        carrera_key=row.get("carrera_key", ""),
        plan_key=row.get("plan_key", ""),
        cod_curso_key=row.get("cod_curso_key", ""),
    )
    resolved_curso.append(curso)
    resolved_carrera.append(car_res)
    resolved_plan.append(plan_res)

si_curlle["curso_resuelto"] = resolved_curso
si_curlle["carrera_resuelta"] = resolved_carrera
si_curlle["plan_resuelto"] = resolved_plan

si_curlle["cod_carrera_out"] = si_curlle["cod_carrera"].fillna("").astype(str).str.strip()
si_curlle["plan_out"] = si_curlle["plan"].fillna("").astype(str).str.strip()

mask_override = si_curlle["curso_resuelto"].fillna("").astype(str).str.strip() != ""
si_curlle.loc[mask_override, "cod_carrera_out"] = si_curlle.loc[mask_override, "carrera_resuelta"].fillna("").astype(str).str.strip()
si_curlle.loc[mask_override, "plan_out"] = si_curlle.loc[mask_override, "plan_resuelto"].fillna("").astype(str).str.strip()

si_curlle_dep = depurar_p03(
    si_df=si_curlle,
    padron_prog_col=padron_prog_col,
    padron_plan_sigu_col=padron_plan_sigu_col,
    padron_cod_col=padron_cod_col,
)

p03_codigo_alumno_from_padron = si_curlle_dep[padron_cod_col].fillna("").astype(str).str.strip()

p03_data = pd.DataFrame({
    "Código Alumno": p03_codigo_alumno_from_padron,
    "Nombre Completo": si_curlle_dep["nombre_completo"],
    "Programa Academico": si_curlle_dep[padron_prog_col].fillna("").astype(str).str.strip() if padron_prog_col else "",
    "Periodo": si_curlle_dep["periodo_fmt"],
    "Cod. Curso": si_curlle_dep["cod_curso"].fillna("").astype(str).map(normalize_cod_curso_spaces),
    "Nom. Curso": si_curlle_dep["curso_resuelto"].fillna("").astype(str).str.strip(),
    "Nota Curlle": si_curlle_dep["nota_curlle"],
    "Plan": si_curlle_dep["plan_out"],
    "Cod. Carrera": si_curlle_dep["cod_carrera_out"],
    "PLAN-SIGU": si_curlle_dep[padron_plan_sigu_col].fillna("").astype(str).str.strip() if padron_plan_sigu_col else "",
})
p03_data["OBS"] = ""

# =========================
# P04
# =========================
programa_series = (
    si_curlle_dep[padron_prog_col].fillna("").astype(str)
    if padron_prog_col and padron_prog_col in si_curlle_dep.columns
    else pd.Series([""] * len(si_curlle_dep), index=si_curlle_dep.index)
)
codigo_escuela_series = programa_series.map(escuela_from_programa).fillna("").astype(str).str.strip()

plan_sigu_series = (
    si_curlle_dep[padron_plan_sigu_col].fillna("").astype(str).str.strip()
    if padron_plan_sigu_col and padron_plan_sigu_col in si_curlle_dep.columns
    else pd.Series([""] * len(si_curlle_dep), index=si_curlle_dep.index)
)

egresado_series = (
    si_curlle_dep["egresado_flag"].fillna(False)
    if "egresado_flag" in si_curlle_dep.columns
    else pd.Series([False] * len(si_curlle_dep), index=si_curlle_dep.index)
)

plan_destino = []
for i in range(len(si_curlle_dep)):
    escuela = str(codigo_escuela_series.iloc[i]).strip()
    plan_sigu = str(plan_sigu_series.iloc[i]).strip()
    egresado = bool(egresado_series.iloc[i])

    if egresado:
        plan_final = plan_sigu
    else:
        plan_final = normalize_plan_by_matrix(plan_sigu, escuela)

    plan_destino.append(plan_final)

plan_destino_series = pd.Series(plan_destino, index=si_curlle_dep.index).fillna("").astype(str).str.strip()

plan_codigo_series = (
    (codigo_escuela_series + "-" + plan_destino_series)
    .where((codigo_escuela_series != "") & (plan_destino_series != ""), "")
)

tipo_matricula_series = si_curlle_dep["cod_curso"].map(tipo_matricula_from_codcurso)

# ✅ Periodo P04:
periodo_p04 = si_curlle_dep["periodo_fmt"].fillna("").astype(str).str.strip().copy()
exsuf_mask = si_curlle_dep["cod_curso"].fillna("").astype(str).map(normalize_cod_curso_spaces).str.startswith("EXSUF")
bad_mask = periodo_p04.isin(["#¿NOMBRE?", "221"])
fix_mask = exsuf_mask | bad_mask

if "periodo_ingreso_fmt" in si_curlle_dep.columns:
    periodo_p04.loc[fix_mask] = si_curlle_dep.loc[fix_mask, "periodo_ingreso_fmt"].fillna("").astype(str).str.strip()

codigo_curso_series = []
for i in range(len(si_curlle_dep)):
    escuela = str(codigo_escuela_series.iloc[i]).strip()
    plan_final = str(plan_destino_series.iloc[i]).strip()
    codcurso = str(si_curlle_dep["cod_curso"].iloc[i]).strip()
    nombre_curso = str(si_curlle_dep["curso_resuelto"].iloc[i]).strip()

    codigo_curso_series.append(
        lookup_codigo_curso_tpa(
            tpa_map=tpa_map,
            escuela=escuela,
            plan=plan_final,
            codcurso=codcurso,
            nombre_curso=nombre_curso,
        )
    )

codigo_curso_series = pd.Series(codigo_curso_series, index=si_curlle_dep.index).fillna("").astype(str).str.strip()

p04_ok_mask = codigo_curso_series.fillna("").astype(str).str.strip().ne("")
p03_data.loc[~p04_ok_mask.values, "OBS"] = "NO EXISTE / NO SIMILAR EN PLAN (NO PASA A P04)"

faltan_codigo_curso = int((~p04_ok_mask).sum())

p04_data = pd.DataFrame({
    "Periodo": periodo_p04,
    "CodigoAlumno": p03_codigo_alumno_from_padron,
    "Seccion": "A",
    "CodigoEscuela": codigo_escuela_series,
    "CodigoPlan": plan_codigo_series,
    "CodigoCurso": codigo_curso_series,
    "TipoMatricula": tipo_matricula_series,
})

p04_data = p04_data.loc[p04_ok_mask.values].copy()

# ✅ EXTRA BLINDAJE ANTI-DUP: por si entra "10A06" repetido desde Curlle
p04_data = p04_data.drop_duplicates(
    subset=["Periodo", "CodigoAlumno", "CodigoCurso", "CodigoPlan", "TipoMatricula"],
    keep="first",
).copy()

# =========================
# ✅ P05 (Carga de notas)
# =========================
nota_series = si_curlle_dep["nota_curlle"].fillna("").astype(str)

p05_data = pd.DataFrame({
    "CodigoAlumno": p03_codigo_alumno_from_padron,
    "Periodo": periodo_p04,
    "CodigoCurso": codigo_curso_series,
    "CodigoPlan": plan_codigo_series,
    "Nota": nota_series,
    "TipoMatricula": tipo_matricula_series,
})

p05_data = p05_data.loc[p04_ok_mask.values].copy()

# ✅ EXTRA BLINDAJE ANTI-DUP
p05_data = p05_data.drop_duplicates(
    subset=["Periodo", "CodigoAlumno", "CodigoCurso", "CodigoPlan", "TipoMatricula", "Nota"],
    keep="first",
).copy()

# =========================
# Render previews + Export
# =========================
def _style_p03(row):
    if str(row.get("OBS", "")).strip() != "":
        return ["background-color: #fff59d"] * len(row)
    return [""] * len(row)


c1, c2, c3 = st.columns(3)
with c1:
    st.subheader("P01 - Existentes en Akademic")
    st.caption(f"Registros: {len(p01_data):,}")
    st.dataframe(p01_data.head(20), use_container_width=True)

with c2:
    st.subheader("P02 - No Akademic / No Curlle")
    st.caption(f"Registros: {len(p02_data):,}")
    st.dataframe(p02_data.head(20), use_container_width=True)

with c3:
    st.subheader("P03 - No Akademic / Sí Curlle (DEPURADO)")
    st.caption(f"Registros: {len(p03_data):,}")
    st.dataframe(p03_data.head(200).style.apply(_style_p03, axis=1), use_container_width=True)

st.divider()
st.subheader("P04 - Matrícula desde P03 (SOLO filas con CodigoCurso válido)")
st.caption(f"Registros: {len(p04_data):,}")
st.dataframe(p04_data.head(50), use_container_width=True)

st.subheader("P05 - Carga de notas (SOLO filas que pasaron a P04)")
st.caption(f"Registros: {len(p05_data):,}")
st.dataframe(p05_data.head(50), use_container_width=True)

missing_tpl = []
for p in [P01_TEMPLATE, P02_TEMPLATE, P03_TEMPLATE, P04_TEMPLATE, P05_TEMPLATE]:
    if not p.exists():
        missing_tpl.append(p.name)

if missing_tpl:
    st.error("Faltan plantillas en /plantillas:\n- " + "\n- ".join(missing_tpl))
    st.stop()

p01_sheet = align_df_to_template_df(P01_TEMPLATE, p01_data)
p02_sheet = align_df_to_template_df(P02_TEMPLATE, p02_data)
p03_sheet = align_df_to_template_df(P03_TEMPLATE, p03_data)

# ✅ FORZAR OBS en el excel descargado (aunque tu plantilla no tenga esa columna)
p03_sheet["OBS"] = p03_data["OBS"].fillna("").astype(str)

p04_sheet = align_df_to_template_df(P04_TEMPLATE, p04_data)
p05_sheet = align_df_to_template_df(P05_TEMPLATE, p05_data)

multi_excel_bytes = build_multi_sheet_excel({
    "P01": p01_sheet,
    "P02": p02_sheet,
    "P03": p03_sheet,
    "P04": p04_sheet,
    "P05": p05_sheet,
})

st.divider()
st.subheader("📦 Descarga")

st.download_button(
    "⬇️ Descargar Excel (P01–P05 en hojas)",
    data=multi_excel_bytes,
    file_name="SALIDAS_REGULARIZACION_AKA.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

if faltan_codigo_curso > 0:
    st.warning(
        f"⚠️ Se EXCLUYERON {faltan_codigo_curso:,} filas de P04 porque no se encontró "
        f"match por nombre/código (ni parecido) en el MISMO plan dentro de TODOS_PLANES_AKADEMIC. "
        f"Revisa P03 columna OBS (se pinta amarillo)."
    )

st.success("Listo ✅ Descargable único generado con hojas P01, P02, P03, P04 y P05.")